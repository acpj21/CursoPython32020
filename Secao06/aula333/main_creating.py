# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Arquivo completo, pois terá várias planilhas em um arquivo
workbook = Workbook()

# Uma planilha específica
# worksheet: Worksheet = workbook.active

# Nome para a Planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)

# Selecionamos a planilha
# worksheet: Worksheet = workbook.active
worksheet: Worksheet = workbook[sheet_name]
# print(workbook.sheetnames)

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     # print(i, student_row)
#     for j, student_column in enumerate(student_row, start=1):
#         # print(i, j, student_column)
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)